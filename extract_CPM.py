import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import zipfile
import io
import os
import re
from math import ceil


def trim_lsc_output_from_stream(content_bytes):
    search_strings = [
        b"\x1bx \x1bt \x1b7\x1bR \x1b2\x12\x1bP\x1bQP\x1bW \x1bH\x1b- \x1bx \x1bR \x1b2\x12\x1bP\x1bQP\x0c",
        b"\x1bx\x00\x1bt\x00\x1b7\x1bR\x00\x1b2\x12\x1bP\x1bQP\x1bW\x00\x1bH\x1b-\x00\x1bx\x00\x1bR\x00\x1b2\x12\x1bP\x1bQP\x0c"
    ]

    last_index = -1
    for search in search_strings:
        idx = content_bytes.rfind(search)
        if idx > last_index:
            last_index = idx + len(search)

    if last_index == -1:
        return "", "❌ No marker string found in the file."

    trimmed_bytes = content_bytes[last_index:]
    trimmed_text = trimmed_bytes.decode("utf-8", errors="ignore")

    patterns = [
        r"\x1bW.", r"\x1bG", r"\x1b-.", r"\x1bH", r"\f"
    ]
    for pat in patterns:
        trimmed_text = re.sub(pat, "", trimmed_text)

    lines = trimmed_text.splitlines()

    filtered_lines = [
        line for line in lines
        if line.strip() and not any(kw in line for kw in ["MISSING SAMPLE", "PAGE:", "INVALID SAMPLE"])
    ]

    header1 = "SAM  POS   TIME    H#        3H    LUMEX  ELAPSED"
    header2 = "NO         MIN             CPM  %ERROR     %      TIME"

    first_header_found = False
    final_lines = []

    i = 0
    while i < len(filtered_lines):
        line = filtered_lines[i]
        if header1 in line:
            if not first_header_found:
                first_header_found = True
                final_lines.append(line)
                if i + 1 < len(filtered_lines) and header2 in filtered_lines[i + 1]:
                    final_lines.append(filtered_lines[i + 1])
                    i += 1
            i += 1
            continue
        if header2 in line and first_header_found:
            i += 1
            continue
        final_lines.append(line)
        i += 1

    cleaned_text = "\n".join(final_lines)
    return cleaned_text, None


def create_and_populate_excel(cleaned_text, excel_path):
    cpm_data = {}
    max_sample_no = 0

    for line in cleaned_text.splitlines():
        parts = re.split(r"\s+", line.strip())
        if len(parts) >= 5 and (parts[1].startswith("**-") or parts[1].startswith("17-")):
            try:
                sample_no = int(parts[0])
                cpm = int(float(parts[4]))
                cpm_data[sample_no] = cpm
                max_sample_no = max(max_sample_no, sample_no)
            except ValueError:
                continue

    wb = Workbook()
    ws = wb.active

    for col in range(1, 19):
        cell = ws.cell(row=1, column=col + 1)
        cell.value = col
        cell.font = Font(bold=True)

    num_rows = ceil(max_sample_no / 18)
    for i in range(num_rows):
        cell = ws.cell(row=i + 2, column=1)
        cell.value = chr(65 + (i % 8))
        cell.font = Font(bold=True)

    for sample_no, cpm in cpm_data.items():
        row_index = ((sample_no - 1) // 18) + 2
        col_index = ((sample_no - 1) % 18) + 2
        ws.cell(row=row_index, column=col_index, value=cpm)

    wb.save(excel_path)


st.title("Extract CPM values from LS6500 output file")

uploaded_file = st.file_uploader("Upload a RECORD.TXT file", type=["txt", "TXT"])

if uploaded_file is not None:
    content_bytes = uploaded_file.read()
    original_content = content_bytes.decode("utf-8", errors="ignore")

    # File naming
    current_date = datetime.now().strftime("%Y%m%d")
    clean_file_name = f"{current_date}_RECORD_clean.txt"
    excel_file_name = f"{current_date}_RESULTS.xlsx"
    original_file_name = f"{current_date}_RECORD.txt"
    zip_file_name = f"{current_date}_RESULTS.zip"

    cleaned_text, error_msg = trim_lsc_output_from_stream(content_bytes)
    if error_msg:
        st.error(error_msg)
    else:
        with open(clean_file_name, "w", encoding="utf-8") as f:
            f.write(cleaned_text)
        create_and_populate_excel(cleaned_text, excel_file_name)

        st.download_button("Download Cleaned Text File", cleaned_text, clean_file_name, mime="text/plain")
        st.download_button("Download Results Excel File", open(excel_file_name, "rb").read(), excel_file_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.download_button("Download Original File", original_content, f"{current_date}_RECORD.txt", mime="text/plain")


        # Check if cleaned_output_path exists
        if not os.path.exists(clean_file_name):
            st.error(f"Cleaned file not found: {clean_file_name}")
        else:
            st.success(f"Cleaned file found: {clean_file_name}")
        
        # Check if excel_output_path exists
        if not os.path.exists(excel_file_name):
            st.error(f"Excel file not found: {excel_file_name}")
        else:
            st.success(f"Excel file found: {excel_file_name}")
  

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as zipf:
            zipf.writestr(f"{current_date}_RECORD.txt", original_content)
            if os.path.exists(clean_file_name):
                zipf.write(clean_file_name, arcname=clean_file_name)
            if os.path.exists(excel_file_name):
                zipf.write(excel_file_name, arcname=excel_file_name)
        zip_buffer.seek(0)

        st.download_button("Download All Outputs (ZIP)", zip_buffer, file_name=zip_file_name, mime="application/zip")
